import datetime
import json
import sys as sys
from flask import render_template, send_file
from typing import Any
from flask import (Flask, Response, app, send_file, render_template,
                   request)
from openpyxl import load_workbook
import requests
from duplex import accumulated_materials, order, toggle_value, calc_quad, calc_duplex, reset_variables, calc_gfci, calc_cutin, calc_surface, calc_duplex_controlled, calc_quad, calc_quad_gfci, calc_quad_cutin, calc_quad_surface, calc_quad_controlled, calc_ff3, calc_ff4, calc_rough_data, calc_cutin_data, calc_lv_switch, calc_hv_switch, calc_hv_dimming, calc_lv_switchCI, calc_hv_switchCI, calc_hv_dimmingCI, calc_wh_120, calc_wh_277, calc_ff3, calc_FC4in, calc_FC6in
import os.path
from google.oauth2 import service_account
from openpyxl.utils import get_column_letter 
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.discovery import build
import json
from openpyxl import Workbook
from list import item_dict



# DEBUG = True
SITE_NAME = "Materialized"

global reordered_materials
items = {}
reordered_materials = {}





app = Flask(__name__)
@app.route('/count.html', methods=['GET', 'POST'])
def count():
    if request.method == 'POST':
        item = request.form['item']
        if item not in items:
            items[item] = 0
    return render_template('count.html', items=items)


# def reorder_materials(materials_dict):
#     reordered_materials = {item: quantity for item, quantity in materials_dict.items() if quantity != 0 and item in item_dict}
#     return reordered_materials



# def reorder_materials(materials_dict):
#     global accumulated_materials
#     reordered_materials = {item: materials_dict.get(item, 0) for item in item_dict}
#     return reordered_materials

@app.route('/calendar.html')
def calendar():
    # Pass the cell values to the template
    return render_template('calendar.html')

@app.route('/index3.html')
def documents():
    # Pass the cell values to the template
    return render_template('index3.html')

@app.route('/increment', methods=['POST'])
def increment():
    item = request.form['item']
    items[item] += 1
    return str(items[item])
  
@app.route('/download')  # Place the decorator before the function definition
def download() -> Response:
    filename = 'templates/old.xlsx'
    return send_file(filename, as_attachment=True)
from flask import jsonify

@app.route('/update-color-mode', methods=['POST'])
def update_color_mode():
    color_mode = request.form.get('color_mode')
    toggle_value(color_mode)  # Pass the color_mode value to duplex.py
    return jsonify({'status': 'success'})  # Return a response to the AJAX request


@app.route('/index.html', methods=['GET', 'POST'])
@app.route('/', methods=['GET', 'POST'], )  # type: ignore
def form() -> Any:
    # global decora, duplex_outlets, gfci, cutins, surface, quad_standard, duplex_controlled, four_square_box, four_square_bracket_box
    # If User hits the Generate button...
    # If Generate button is not pressed, just show index.html
    global form_data
    global item_dict
    global accumulated_materials
    global reordered_materials
    if request.method == 'POST':
        reset_variables()
        reset_dict()
        
        # outlet = request.form.get('outlet_type')  # Retrieve the value of the toggle switch
        # toggle_value(outlet)
        
        duplex_outlets = (request.form['standard'])
        if duplex_outlets:
            duplex_outlets = int(duplex_outlets)
            form_data['duplex_outlets'] = int(request.form.get('standard', 0))
        else:
            duplex_outlets  = 0  
            form_data['duplex_outlets'] = 0

                        
        # decora = (request.form['decora'])
        # if decora:
        #     decora = int(decora)
        #     form_data['decora'] = int(request.form.get('decora', 0))
        # else:
        #     decora  = 0  
        #     form_data['decora'] = 0
            
        gfci = (request.form['gfci'])
        if gfci:
            gfci = int(gfci)
            form_data['gfci'] = int(request.form.get('gfci', 0))
        else:
            gfci  = 0
            
        cutin = (request.form['cutin'])
        if cutin:
            cutin = int(cutin)
            form_data['cutin'] = int(request.form.get('cutin', 0))
        else:
            cutin  = 0
            
        surface = (request.form['surface'])
        if surface:
            surface = int(surface)  
            form_data['surface'] = int(request.form.get('surface', 0))
        else:
            surface  = 0
            
        duplex_controlled = (request.form['1switch'])
        if duplex_controlled:
            duplex_controlled = int(duplex_controlled)
            form_data['duplex_controlled'] = int(request.form.get('1switch', 0))
        else:
            duplex_controlled  = 0
        
        quad_standard = (request.form['quad_standard'])
        if quad_standard:
            quad_standard = int(quad_standard)
            form_data['quad_standard'] = int(request.form.get('quad_standard', 0))
        else:
            quad_standard  = 0  
    
        # quad_decora = (request.form['quad_decora'])
        # if quad_decora:
        #     quad_decora = int(quad_decora)
        #     form_data['quad_decora'] = int(request.form.get('quad_decora', 0))
        # else:
        #     quad_decora  = 0
            
        quad_gfci = (request.form['quad_gfci'])
        if quad_gfci:
            quad_gfci = int(quad_gfci)
            form_data['quad_gfci'] = int(request.form.get('quad_gfci', 0))
        else:
            quad_gfci  = 0
            
        quad_cutin = (request.form['quad_cutin'])
        if quad_cutin:
            quad_cutin = int(quad_cutin)
            form_data['quad_cutin'] = int(request.form.get('quad_cutin', 0))
        else:
            quad_cutin  = 0
            
        quad_surface = (request.form['quad_surface'])
        if quad_surface:
            quad_surface = int(quad_surface)
            form_data['quad_surface'] = int(request.form.get('quad_surface', 0))
        else:
            quad_surface  = 0
            
        quad_controlled = (request.form['quad_controlled'])
        if quad_controlled:
            quad_controlled = int(quad_controlled)
            form_data['quad_controlled'] = int(request.form.get('quad_controlled', 0))
        else:
            quad_controlled  = 0

        ff3 = (request.form['3-wire'])
        if ff3:
            ff3 = int(ff3)
            form_data['ff3'] = int(request.form.get('3-wire', 0))
        else:
            ff3  = 0
            
        ff4 = (request.form['4-wire'])
        if ff4:
            ff4 = int(ff4)
            form_data['ff4'] = int(request.form.get('4-wire', 0))
        else:
            ff4  = 0

        rough_data = (request.form['rough_in_data'])
        if rough_data:
            rough_data = int(rough_data)
            form_data['rough_data'] = int(request.form.get('rough_in_data', 0))
        else:
            rough_data  = 0
            
        cutin_data = (request.form['cutin_data'])
        if cutin_data:
            cutin_data = int(cutin_data)
            form_data['cutin_data'] = int(request.form.get('cutin_data', 0))
        else:
            cutin_data  = 0

        lv_switch = (request.form['lv_switch'])
        if lv_switch:
            lv_switch = int(lv_switch)
            form_data['lv_switch'] = int(request.form.get('lv_switch', 0))
        else:
            lv_switch  = 0
            
        hv_switch = (request.form['hv_switch'])
        if hv_switch:
            hv_switch = int(hv_switch)
            form_data['hv_switch'] = int(request.form.get('hv_switch', 0))
        else:
            hv_switch  = 0
            
        hv_dimming = (request.form['hv_dimming'])
        if hv_dimming:
            hv_dimming = int(hv_dimming)
            form_data['hv_dimming'] = int(request.form.get('hv_dimming', 0))
        else:
            hv_dimming  = 0
            
        hv_dimmingCI = (request.form['hv_dimmingCI'])
        if hv_dimmingCI:
            hv_dimmingCI = int(hv_dimmingCI)
            form_data['hv_dimmingCI'] = int(request.form.get('hv_dimmingCI', 0))
        else:
            hv_dimmingCI = 0
            
        hv_switchCI = (request.form['hv_switchCI'])
        if hv_switchCI:
            hv_switchCI = int(hv_switchCI)
            form_data['hv_switchCI'] = int(request.form.get('hv_switchCI', 0))
        else:
            hv_switchCI = 0
            
        lv_switchCI = (request.form['lv_switchCI'])
        if lv_switchCI:
            lv_switchCI = int(lv_switchCI)
            form_data['lv_switchCI'] = int(request.form.get('lv_switchCI', 0))
        else:
            lv_switchCI = 0

        wh_120 = (request.form['wh_120'])
        if wh_120:
            wh_120 = int(wh_120)
            form_data['wh_120'] = int(request.form.get('wh_120', 0))
        else:
            wh_120  = 0

        wh_277 = (request.form['wh_277'])
        if wh_277:
            wh_277 = int(wh_277)
            form_data['wh_277'] = int(request.form.get('wh_277', 0))
        else:
            wh_277  = 0
            
        FC4in = (request.form['FC4in'])
        if FC4in:
            FC4in = int(FC4in)
            form_data['FC4in'] = int(request.form.get('FC4in', 0))
        else:
            FC4in  = 0

        FC6in = (request.form['FC6in'])
        if FC6in:
            FC6in = int(FC6in)
            form_data['FC6in'] = int(request.form.get('FC6in', 0))
        else:
            FC6in  = 0
        workbook = Workbook()
        sheet = workbook.active

        # Write form data to the worksheet
        
        sheet['E5'] = duplex_outlets
        sheet['E6'] = gfci
        sheet['E7'] = cutin
        sheet['E8'] = surface
        sheet['E9'] = duplex_controlled
        
        sheet['E12'] = quad_standard
        sheet['E13'] = quad_gfci                                                                                                                                                                
        sheet['E14'] = quad_cutin
        sheet['E15'] = quad_surface
        sheet['E16'] = quad_controlled 
                                        
        sheet['E19'] = lv_switch
        sheet['E20'] = hv_dimmingCI
        sheet['E21'] = hv_switch
        sheet['E22'] = hv_switchCI
        sheet['E23'] = hv_dimming
        sheet['E24'] = lv_switchCI
              
        sheet['E27'] = rough_data
        sheet['E28'] = cutin_data                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
        sheet['E29'] = ff3
        sheet['E30'] = ff4
        sheet['E31'] = wh_120
        sheet['E32'] = wh_277
        sheet['E33'] = FC4in
        sheet['E34'] = FC6in
        
        # need to get rid of standard/decora counts and add a toggle so that the total plug count can include cut-in and surface mounted with the correct type of outlet
        sheet['C5'] = "Duplex Bracket Box"
        sheet['C6'] = "GFCI"
        sheet['C7'] = "Cut-in"
        sheet['C8'] = "Surface Mounted"
        sheet['C9'] = "Controlled"
        # need to add sub catagory of single outlet vs whole outlet controlled(subtract amount of controlled from whole plug count)
        
        # need to get rid of standard/decora counts and refer to toggle above so that the total plug count can include cut-in and surface mounted with the correct type of outlet
        sheet['C12'] = "Quad Bracket Box"
        sheet['C13'] = "GFCI"                                                                                                                                                               
        sheet['C14'] = "Cut-in"
        sheet['C15'] = "Surface Mounted"
        sheet['C16'] = "Controlled"
        # need to add sub catagory of single outlet vs whole outlet controlled(subtract amount of controlled from whole plug count)

                          
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          
        sheet['C19'] = "Low-Voltage"
        sheet['C20'] = "Low-Voltage Cut-in"
        sheet['C21'] = "Line-Voltage"
        sheet['C22'] = "Line-Voltage Cut-in"
        sheet['C23'] = "Line-Volt Dimming"
        sheet['C24'] = "Line-Volt Dimming Cut-in"
        
        sheet['C27'] = "Rough-in Data"
        sheet['C28'] = "Cut-in Data"
        sheet['C29'] = "3-wire Furniture Feed"
        sheet['C30'] = "4-wire Funiture Feed"
        sheet['C31'] = "208V 40A Instahot"
        sheet['C32'] = "277V 30A Instahot"
        sheet['C33'] = "6in Floor Device"
        sheet['C34'] = "4in Floor Device"


#  Save the workbook as an XLSX file
        workbook.save('form_data.xlsx')
        calc_duplex(duplex_outlets) 
       # calc_decora(decora)
        calc_gfci(gfci)
        calc_cutin(cutin)
        calc_surface(surface)
        calc_duplex_controlled(duplex_controlled)
    
        calc_quad(quad_standard)
        # calc_quad_decora(quad_decora)
        calc_quad_gfci(quad_gfci)
        calc_quad_cutin(quad_cutin)
        calc_quad_surface(quad_surface)
        calc_quad_controlled(quad_controlled)
        
        calc_ff3(ff3)
        calc_ff4(ff4)
        calc_rough_data(rough_data)
        calc_cutin_data(cutin_data)
        calc_lv_switch(lv_switch)
        calc_hv_switch(hv_switch)
        calc_hv_dimming(hv_dimming)
        calc_lv_switchCI(lv_switchCI)
        calc_hv_switchCI(hv_switchCI)
        calc_hv_dimmingCI(hv_dimmingCI)
        calc_wh_120(wh_120)
        calc_wh_277(wh_277) 
        calc_FC4in(FC4in)
        calc_FC6in(FC6in) 
# Reorder the accumulated_materials dictionary
 # Reorder the accumulated_materials dictionary
#         reordered_materials = reorder_materials(accumulated_materials)

# # Assign the reordered dictionary back to accumulated_materials
#         accumulated_materials = reordered_materials
        sorted_dict = order()

        filename = 'templates/matlist.xlsx'
        export_to_excel(accumulated_materials, filename)

        return render_template('result.html', materials=sorted_dict, form_data=form_data),reset_variables()
        #Redirect to result page
        # return redirect(url_for('result'))  # type: ignore
    # show html page
    return render_template('index.html')  # type: ignore







from openpyxl import Workbook
from openpyxl.utils import get_column_letter
def export_to_excel(data_dict, filename):
    wb = load_workbook('templates/matlist.xlsx')
    ws = wb.active

    # Write headers
    headers = ['Value', 'Key']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Write data
    row_num = 12
    for key, value in data_dict.items():
        col_num = 3
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}{row_num}"] = value

        col_num = 4
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}{row_num}"] = key

        row_num += 1

    # Save workbook
    wb.save('templates/old.xlsx')

        
def reset_dict():
    global form_data
    form_data = {}
    
    
@app.route('/todo.html')
def todo() -> str:   # show the form, it wasn't submitted
    return render_template('todo.html')
@app.route('/upload-to-sheets', methods=['POST'])
def upload_to_sheets():
    # Google Sheets API credentials
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    spreadsheet_id = '*****ENTER SPREADSHEET ID'
    range_xlsx = '# of Installed Devices!A1:Z100'  # Range for XLSX data
    range_html = 'Walls Material List!C13:F1500'  # Range for HTML table data
    secret_file = os.path.join(os.getcwd(), 'credentials.json') *****SAVE GOOGLE API SERVICE ACCOUNT INFO IN ROOT FOLDER NAME credentials.json*****

    # Load the XLSX file
    xlsx_file_path = "form_data.xlsx"
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    # Get all the values from the sheet
    data_xlsx = sheet.iter_rows(values_only=True)
    # Assuming the first row contains column headers, skip it
    headers_xlsx = next(data_xlsx)
    rows_xlsx = list(data_xlsx)

    # Convert datetime objects to strings
    rows_xlsx = [[str(cell) if isinstance(cell, datetime.datetime) else cell for cell in row] for row in rows_xlsx]

    # Parse the HTML table data from the POST request
    data_html = json.loads(request.data)
    headers_html = data_html[0]
    rows_html = data_html[1:]

    # Swap the sides of columns in HTML table data
    rows_html_swapped = [[row[i] for i in range(len(row)-1, -1, -1)] for row in rows_html]
    # Authenticate with Google Sheets API
    creds = service_account.Credentials.from_service_account_file(secret_file, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)

    try:
        # Prepare the XLSX data for uploading to Google Sheets
        values_xlsx = {'values': [headers_xlsx] + rows_xlsx}

        # Call the Google Sheets API to update the spreadsheet with the XLSX data
        result_xlsx = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range_xlsx,
            valueInputOption='USER_ENTERED', body=values_xlsx
        ).execute()

        print(f"XLSX data uploaded to Google Sheets: {result_xlsx.get('updatedCells')} cells updated")

        # Prepare the HTML table data with swapped columns for uploading to Google Sheets
        values_html = {'values': rows_html_swapped}

        # Call the Google Sheets API to update the spreadsheet with the HTML table data
        result_html = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range_html,
            valueInputOption='USER_ENTERED', body=values_html
        ).execute()

        print(f"HTML table data uploaded to Google Sheets: {result_html.get('updatedCells')} cells updated")

        return "Data uploaded to Google Sheets successfully!"
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return "Error occurred while uploading data to Google Sheets"


@app.route('/result.html')
def result():
    # Pass the cell values to the template
    return render_template('result.html')
    # return   send_file(filename, as_attachment=True, ) # type: ignore


@app.route('/manifest.json')  # type: ignore
def manifest() -> Response:
    return app.send_static_file('manifest.json')

# Serve service worker file


@app.route('/sw.js')
def service_worker() -> Response:
    return app.send_static_file('sw.js')

# Cache static assets


@app.after_request
def add_header(response) -> Response:
    response.headers['Cache-Control'] = 'static, max-age=31536000'
    return response


if __name__ == '__main__':
    app.run(debug = False, port = 5050, host = '0.0.0.0')
